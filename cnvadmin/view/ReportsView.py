from datetime import datetime, timedelta, time
from django.shortcuts import render
from firebase_admin import firestore
import pytz
from django.utils.html import mark_safe
from django.template.defaultfilters import escapejs
from ..config import PG_HOST, PG_PORT, PG_DB, PG_USER, PG_PASSWORD
from ..config import MAX_DISTANCE
from django.utils.safestring import mark_safe
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from cnvadmin.dto.dto import customerDto, customerFilterDto, newCustomerDto, newCustomerFilterDto, visitDto, visitFilterDto
from django.http import HttpResponse
from ..utils.queryFieldsPointerUtils import CUSTOMERQUERY_CUSTOMER_ID, CUSTOMERQUERY_CUSTOMER_ADDRESS, CUSTOMERQUERY_CUSTOMER_CONTACT_PERSON, CUSTOMERQUERY_CUSTOMER_NAME, CUSTOMERQUERY_CUSTOMER_NO_TELP, CUSTOMERQUERY_CUSTOMER_DAERAH, CUSTOMERQUERY_MASTER_LIST_NAME, CUSTOMERQUERY_SALES_ORDER_ID, CUSTOMERQUERY_SALES_ORDER_DATE, CUSTOMERQUERY_SALES_ORDER_DELIVERY_DATE, CUSTOMERQUERY_SALES_ORDER_INVOICE_ID, CUSTOMERQUERY_SALES_ORDER_INVOICE_DATE, CUSTOMERQUERY_SALES_ORDER_INVOICE_DUE_DATE, CUSTOMERQUERY_SALES_ORDER_INVOICE_DATE_COMPLETED, CUSTOMERQUERY_SALES_ORDER_TOTAL, CUSTOMERQUERY_SALES_ORDER_STATUS_ID, CUSTOMERQUERY_PAYMENT_METHOD_NAME, CUSTOMERQUERY_SALES_ORDER_INVOICE_INVOICE_NUMBER, CUSTOMERQUERY_CUSTOMER_REFERENCE_CODE, CUSTOMERQUERY_APPROVAL_STATUS_NAME, CUSTOMERQUERY_TOTAL_PAYMENT_AMOUNT
from ..utils.queryFieldsPointerUtils import ITEMQUERY_SALES_ORDER_DETAIL_SALES_ORDER_ID, ITEMQUERY_SALES_ORDER_DETAIL_ID, ITEMQUERY_ITEM_NAME, ITEMQUERY_SALES_ORDER_DETAIL_QUANTITY, ITEMQUERY_SALES_ORDER_DETAIL_PRICE
from ..utils.queryFieldsPointerUtils import DEBTQUERY_CUSTOMER_ID, DEBTQUERY_SALES_ORDER_ID, DEBTQUERY_SALES_ORDER_DATE, DEBTQUERY_SALES_ORDER_DELIVERY_DATE, DEBTQUERY_SALES_ORDER_INVOICE_ID, DEBTQUERY_SALES_ORDER_INVOICE_DATE, DEBTQUERY_SALES_ORDER_INVOICE_DUE_DATE, DEBTQUERY_SALES_ORDER_INVOICE_TOTAL, DEBTQUERY_PAYMENT_METHOD_NAME, DEBTQUERY_SALES_ORDER_INVOICE_INVOICE_NUMBER, DEBTQUERY_TOTAL_PAYMENT_AMOUNT, DEBTQUERY_DEBT
from ..utils.queryFieldsPointerUtils import PAYMENTDETAILSQUERY_CUSTOMER_ID, PAYMENTDETAILSQUERY_SALES_ORDER_ID, PAYMENTDETAILSQUERY_SALES_ORDER_INVOICE_ID, PAYMENTDETAILSQUERY_SALES_ORDER_INVOICE_DATE, PAYMENTDETAILSQUERY_SALES_ORDER_PAYMENT_DATE, PAYMENTDETAILSQUERY_SALES_ORDER_PAYMENT_AMOUNT, PAYMENTDETAILSQUERY_SALES_ORDER_PAYMENT_DESCRIPTION, PAYMENTDETAILSQUERY_TOTAL_PAYMENT_AMOUNT
import psycopg2
from math import radians, sin, cos, sqrt, atan2

# Create a Firestore client
db = firestore.client()

def reports(request):
    if request.method == 'POST':
        # Get the selected date from the form
        valueDateFrom = request.POST.get('dateFromFilter')
        valueDateTo = request.POST.get('dateToFilter')

        selected_dateFrom = datetime.strptime(valueDateFrom, '%Y-%m-%d')
        selected_dateTo = datetime.strptime(valueDateTo, '%Y-%m-%d') + timedelta(days=1)
        reportType = request.POST.get('reportTypeFilter')
        
        query = db.collection('visits').where('date', '>=', selected_dateFrom).where('date', '<', selected_dateTo)
        docs = query.stream()
        docs = filter(lambda x: x.to_dict().get('user_id') != 'wjb.canvasser@gmail.com', docs)
        #if (reportType == 'customer'):
        #    docs = sorted(docs, key=lambda x: (x.to_dict().get('area'), x.to_dict().get('customerCode'), x.to_dict().get('CheckInTime')))
        #else:
        #    docs = sorted(docs, key=lambda x: (x.to_dict().get('user_id'), x.to_dict().get('CheckInTime')))

        #query = db.collection('visits').where('date', '>=', selected_dateFrom).where('date', '<', selected_dateTo)
        #docs = query.stream()
        #docs = filter(lambda x: x.to_dict().get('user_id') != 'wjb.canvasser@gmail.com', docs)

        workbook = Workbook()
        worksheet = workbook.active

        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

        row_num = 1
        if reportType == 'canvasser':
            title = f"Canvasser Reports {valueDateFrom} - {valueDateTo}"
        else:
            title = f"Customer Visit Reports {valueDateFrom} - {valueDateTo}"

        col_num = 1
        col_letter = get_column_letter(col_num)
        worksheet[f"{col_letter}{row_num}"] = title

        if reportType == 'canvasser':
            header = ['Canvasser','Date','Master List','Name','Address','Daerah','Last Invoice','Last DueDate','Last Delivery','Last Complete','Last Order','Payment Type','Payment Details','Debt','Raport','Note','Check-In','Distance','Check-Out','Distance','Visit Time','Prev Loc Time', 'Skip','Location']
        else:
            header = ['Master List','Customer Code','Name','Address','Daerah','Last Invoice','Last DueDate','Last Delivery','Last Complete','Last Order','Payment Type','Payment Details','Debt','Raport','Note','Date','Canvasser','Check-In','Distance','Check-Out','Distance','Visit Time','Prev Loc Time', 'Skip','Location']

        row_num += 1
        for col_num, cell_value in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            worksheet[f"{col_letter}{row_num}"] = cell_value

        visits = []
        for doc in docs:
            docDict = doc.to_dict()
            location = ''
            if ('location' in docDict):
                location = doc.get('location')

            visit = visitDto(
                id = doc.id,
                customerCode = int(doc.get('customerCode')),
                address = doc.get('address'),      
                area = doc.get('area'),
                checkInLocation = doc.get('checkInLocation'),
                checkInTime = doc.get('checkInTime'),
                checkInDistance= 0,
                checkOutLocation = doc.get('checkOutLocation'),
                checkOutTime = doc.get('checkOutTime'),
                checkOutDistance=0,
                isSkip= False,
                skipTime = '',
                daerah='',
                date = doc.get('date'),
                isCheckIn = doc.get('isCheckIn'),
                isCheckOut = doc.get('isCheckOut'),
                note = doc.get('note'),
                phone = doc.get('phone'),
                pic = doc.get('pic'),
                user_id = doc.get('user_id'),
                last_complete= '',
                last_duedate= '',
                last_invoice= '',
                raport= '',
                last_order='',
                last_delivery='',
                last_payment='',
                last_count_days='',
                debt='',
                payment_details='',
                master_list_order=doc.get('master_list_order'),
                location=location
            )
            if "isSkip" in doc.to_dict():
                visit.isSkip=doc.get('isSkip')
            if "skipTime" in doc.to_dict():
                visit.skipTime=doc.get('skipTime')
            if "daerah" in doc.to_dict():
                visit.daerah=doc.get('daerah')
            if "last_complete" in doc.to_dict():
                visit.last_complete=doc.get('last_complete')
            if "last_complete" in doc.to_dict():
                visit.last_complete=doc.get('last_complete')
            if "last_duedate" in doc.to_dict():
                visit.last_duedate=doc.get('last_duedate')
            if "last_invoice" in doc.to_dict():
                visit.last_invoice=doc.get('last_invoice')
            if "raport" in doc.to_dict():
                visit.raport=doc.get('raport')
            if "last_order" in doc.to_dict():
                visit.last_order=doc.get('last_order')
            if "last_delivery" in doc.to_dict():
                visit.last_delivery=doc.get('last_delivery')
            if "last_payment" in doc.to_dict():
                visit.last_payment=doc.get('last_payment')
            if "last_count_days" in doc.to_dict():
                visit.last_count_days=doc.get('last_count_days')
            if "debt" in doc.to_dict():
                visit.debt=doc.get('debt')
            if "payment_details" in doc.to_dict():
                visit.payment_details=doc.get('payment_details')
            if doc.get('checkInLocation') != '' and doc.get('location') != '':
                aPos = doc.get('location').split(',')
                bPos = doc.get('checkInLocation').split(',')
                visit.checkInDistance = round(calc_dist(float(aPos[0]), float(aPos[1]), float(bPos[0]), float(bPos[1])),2)
            if doc.get('checkOutLocation') != '' and doc.get('location') != '':
                aPos = doc.get('location').split(',')
                bPos = doc.get('checkOutLocation').split(',')
                visit.checkOutDistance = round(calc_dist(float(aPos[0]), float(aPos[1]), float(bPos[0]), float(bPos[1])),2)
            
            visits.append(visit)

        if reportType == 'customer':
            lastCustomer = 'temp'
            sorted_visits = sorted(visits, key=custom_sort_key)
            
            for visit in sorted_visits:
                checkInIime = visit.checkInTime
                if checkInIime:
                    checkInTime = (checkInIime + timedelta(hours=7)).replace(tzinfo=None)
                else:
                    checkInTime = None

                checkOutIime = visit.checkOutTime
                if checkOutIime:
                    checkOutTime = (checkOutIime + timedelta(hours=7)).replace(tzinfo=None)
                else:
                    checkOutTime = None

                if (lastCustomer != visit.customerCode):                        
                    lastCheckOut = None                
                    lastCustomer = visit.customerCode
                    visitTime = ''
                    prevLocTime = ''

                #insert data
                row_num += 1              
                if checkInTime is not None and checkOutTime is not None:
                    visitTimeToCalculate = checkOutTime - checkInTime          
                    visitTimeSeconds = visitTimeToCalculate.total_seconds()
                    hours, remainder = divmod(visitTimeSeconds, 3600)
                    minutes, seconds = divmod(remainder, 60)
                    hours = int(hours)
                    minutes = int(minutes)
                    seconds = int(seconds)
                    visitTime = f"{hours}h {minutes}m {seconds}s"
                else:
                    visitTime = ''

                if lastCheckOut is not None:
                    if checkInTime is not None:
                        prevLocTimeToCalculate = checkInTime.date() - lastCheckOut.date()
                        days = abs(prevLocTimeToCalculate.days)
                        prevLocTime = f"{days} day(s)"
                    else:
                        prevLocTime = ''
                else:
                    prevLocTime = ''
               
                data = [visit.master_list_order, visit.customerCode, visit.pic, visit.address, visit.daerah, visit.last_invoice, visit.last_duedate, visit.last_delivery, visit.last_complete, visit.last_order, visit.last_payment, visit.payment_details, visit.debt, visit.raport, visit.note, visit.date.strftime("%Y-%m-%d"), visit.user_id, checkInTime, visit.checkInDistance, checkOutTime, visit.checkOutDistance, visitTime, prevLocTime, visit.isSkip, visit.location ]

                lastCheckOut = checkOutTime

                for col_num, cell_value in enumerate(data, 1):
                    col_letter = get_column_letter(col_num)
                    worksheet[f"{col_letter}{row_num}"] = cell_value

                if visit.checkInDistance > MAX_DISTANCE:
                    col_letter = get_column_letter(19)
                    worksheet[f"{col_letter}{row_num}"].fill = red_fill
                if visit.checkOutDistance > MAX_DISTANCE:
                    col_letter = get_column_letter(21)
                    worksheet[f"{col_letter}{row_num}"].fill = red_fill

        else:
            currentUser = 'temp'
            canvasserList = []
            totalVisits = {}
            totalSkip = {}
            totalCheck = {}
            totalIn = {}
            totalOut = {}
            totalInRed = {}
            totalInGreen = {}
            totalOutRed = {}
            totalOutGreen = {}

            sorted_visits = sorted(visits, key=custom_sort_key_user)
            for visit in sorted_visits:
                if (currentUser != visit.user_id):                        
                    canvasserList.append(visit.user_id)
                    
                    totalVisits[visit.user_id] = 0
                    totalSkip[visit.user_id] = 0
                    totalCheck[visit.user_id] = 0
                    totalIn[visit.user_id] = 0
                    totalOut[visit.user_id] = 0
                    totalInRed[visit.user_id] = 0
                    totalInGreen[visit.user_id] = 0
                    totalOutRed[visit.user_id] = 0
                    totalOutGreen[visit.user_id] = 0

                    lastCheckOut = None                
                    currentUser = visit.user_id
                    lastDate = None
                    visitTime = ''
                    prevLocTime = ''

                totalVisits[visit.user_id] = totalVisits[visit.user_id] + 1

                if visit.isSkip is True:
                    totalSkip[visit.user_id] += 1
                
                checkInIime = visit.checkInTime
                if checkInIime:
                    checkInTime = (checkInIime + timedelta(hours=7)).replace(tzinfo=None)
                    totalIn[visit.user_id] = totalIn[visit.user_id] + 1
                else:
                    checkInTime = None

                checkOutIime = visit.checkOutTime
                if checkOutIime:
                    checkOutTime = (checkOutIime + timedelta(hours=7)).replace(tzinfo=None)
                    totalOut[visit.user_id] = totalOut[visit.user_id] + 1
                else:
                    checkOutTime = None
                
                if checkInTime and checkOutTime:
                    totalCheck[visit.user_id] = totalCheck[visit.user_id] + 1

                if lastDate != visit.date.strftime("%Y-%m-%d"):
                    lastDate = None

                #insert data
                row_num += 1              
                if checkInTime is not None and checkOutTime is not None:
                    visitTimeToCalculate = checkOutTime - checkInTime          
                    visitTimeSeconds = visitTimeToCalculate.total_seconds()
                    hours, remainder = divmod(visitTimeSeconds, 3600)
                    minutes, seconds = divmod(remainder, 60)
                    hours = int(hours)
                    minutes = int(minutes)
                    seconds = int(seconds)
                    visitTime = f"{hours}h {minutes}m {seconds}s"
                else:
                    visitTime = ''

                if lastCheckOut is not None:
                    if checkInTime is not None and lastDate is not None:
                        prevLocTimeToCalculate = checkInTime - lastCheckOut
                        prevLocTimeSeconds = prevLocTimeToCalculate.total_seconds()
                        hours, remainder = divmod(prevLocTimeSeconds, 3600)
                        minutes, seconds = divmod(remainder, 60)
                        hours = int(hours)
                        minutes = int(minutes)
                        seconds = int(seconds)
                        prevLocTime = f"{hours}h {minutes}m {seconds}s"
                    else:
                        prevLocTime = ''
                else:
                    prevLocTime = ''

                lastDate = visit.date.strftime("%Y-%m-%d")
                lastCheckOut = checkOutTime
                
                data = [visit.user_id, visit.date.strftime("%Y-%m-%d"), visit.master_list_order, visit.pic, visit.address, visit.daerah, visit.last_invoice, visit.last_duedate, visit.last_delivery, visit.last_complete, visit.last_order, visit.last_payment, visit.payment_details, visit.debt, visit.raport, visit.note, checkInTime, visit.checkInDistance, checkOutTime, visit.checkOutDistance, visitTime, prevLocTime, visit.isSkip, visit.location ]

                for col_num, cell_value in enumerate(data, 1):
                    col_letter = get_column_letter(col_num)
                    worksheet[f"{col_letter}{row_num}"] = cell_value
                
                if visit.checkInDistance > MAX_DISTANCE:
                    col_letter = get_column_letter(18)
                    worksheet[f"{col_letter}{row_num}"].fill = red_fill
                    totalInRed[visit.user_id] = totalInRed[visit.user_id] + 1
                elif checkInTime:
                    totalInGreen[visit.user_id] = totalInGreen[visit.user_id] + 1
                    
                if visit.checkOutDistance > MAX_DISTANCE:
                    col_letter = get_column_letter(20)
                    worksheet[f"{col_letter}{row_num}"].fill = red_fill
                    totalOutRed[visit.user_id] = totalOutRed[visit.user_id] + 1
                elif checkOutTime:
                    totalOutGreen[visit.user_id] = totalOutGreen[visit.user_id] + 1

            workbook.create_sheet("recap")
            worksheet = workbook["recap"]
            row_num = 1
            data = ["Canvasser","Visit Assigned","Complete Check","Percent","CheckIn","Percent","Green","Percent","Red","Percent","CheckOut","Percent","Green","Percent","Red","Percent","Skip","Percent"]
            for col_num, cell_value in enumerate(data, 1):
                col_letter = get_column_letter(col_num)
                worksheet[f"{col_letter}{row_num}"] = cell_value

            for name in canvasserList:
                row_num += 1
                data = [name, totalVisits[name], totalCheck[name], "{:.2f}%".format(totalCheck[name] / totalVisits[name] * 100) if totalVisits[name] != 0 else "N/A", totalIn[name], "{:.2f}%".format(totalIn[name] / totalVisits[name] * 100) if totalVisits[name] != 0 else "N/A",totalInGreen[name], "{:.2f}%".format(totalInGreen[name] / totalIn[name] * 100) if totalIn[name] != 0 else "N/A",totalInRed[name], "{:.2f}%".format(totalInRed[name] / totalIn[name] * 100) if totalIn[name] != 0 else "N/A",totalOut[name], "{:.2f}%".format(totalOut[name] / totalVisits[name] * 100) if totalVisits[name] != 0 else "N/A",totalOutGreen[name], "{:.2f}%".format(totalOutGreen[name] / totalOut[name] * 100) if totalOut[name] != 0 else "N/A",totalOutRed[name], "{:.2f}%".format(totalOutRed[name] / totalOut[name] * 100) if totalOut[name] != 0 else "N/A",totalSkip[name], "{:.2f}%".format(totalSkip[name] / totalVisits[name] * 100) if totalVisits[name] != 0 else "N/A"]
                for col_num, cell_value in enumerate(data, 1):
                    col_letter = get_column_letter(col_num)
                    worksheet[f"{col_letter}{row_num}"] = cell_value

        # Create a response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="data.xlsx"'

        # Save the workbook to the response
        workbook.save(response)

        return response

    else:
        return render(request, 'cnvadmin/reports.html')
    
def reports2(request):
    if request.method == 'POST':
        # Get the selected date from the form
            # Connect to your postgres DB
        conn = psycopg2.connect(
            host=PG_HOST,
            port=PG_PORT,
            database=PG_DB,
            user=PG_USER,
            password=PG_PASSWORD
        )
        # Open a cursor to perform database operations
        cur = conn.cursor()

        locationDocs = db.collection('locations').stream()
        locationDataDict = {}

        for locationDoc in locationDocs:
            locationData = locationDoc.to_dict()
            customerCode = locationData.get('customerCode')
            if customerCode is not None:
                if customerCode not in locationDataDict:
                    locationDataDict[customerCode] = []
                    locationDataDict[customerCode].append(locationData)

        reportType = request.POST.get('reportTypeFilter')

        workbook = Workbook()
        worksheet = workbook.active

        row_num = 1
        if reportType == 'location':
            title = f"Customer Location Report"

        col_num = 1
        col_letter = get_column_letter(col_num)
        worksheet[f"{col_letter}{row_num}"] = title

        if reportType == 'location':
            header = ['Master List','Customer Code','Name','Address','Last Invoice','Last DueDate','Last Delivery','Last Complete','Last Order','Payment Type','Payment Details','Debt','Raport','Location']

        row_num += 1
        for col_num, cell_value in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            worksheet[f"{col_letter}{row_num}"] = cell_value

        customerQuery = """
        WITH RankedSalesOrders AS (
            SELECT
                customer.id as customer_id,
                customer.address as customer_address, 
                customer.contact_person as customer_contact_person,
                customer.name as customer_name,
                customer.no_telp as customer_no_telp,
                customer.daerah as customer_daerah,
                master_list.name as master_list_name,
                sales_order.id as sales_order_id,
                sales_order.date as sales_order_date,
                sales_order.delivery_date as sales_order_delivery_date,
                sales_order_invoice.id as sales_order_invoice_id,
                sales_order_invoice.date as sales_order_invoice_date,
                sales_order_invoice.due_date as sales_order_invoice_due_date,
                sales_order_invoice.date_completed as sales_order_invoice_date_completed,
                COALESCE(sales_order_invoice.total,0) as sales_order_invoice_total,
                sales_order.status_id as sales_order_status_id,
                approval_status.name as approval_status_name,
                payment_method.name as payment_method_name, 
                sales_order_invoice.invoice_number as sales_order_invoice_invoice_number,
                customer.reference_code as customer_reference_code,
                SUM(sales_order_payment.amount) AS total_payment_amount,
                ROW_NUMBER() OVER (PARTITION BY customer.id ORDER BY sales_order_invoice.date DESC) AS rn
            FROM
                customer
                LEFT JOIN master_list on (customer.master_list_id = master_list.id)
                LEFT JOIN sales_order ON (customer.id = sales_order.customer_id)
                LEFT JOIN sales_order_invoice ON (sales_order.id = sales_order_invoice.sales_order_id)            
                LEFT JOIN approval_status ON (approval_status.id = sales_order.status_id)
                LEFT JOIN payment_method ON (sales_order_invoice.payment_method_id = payment_method.id)
                LEFT JOIN sales_order_invoice_payment_details ON (sales_order_invoice.id = sales_order_invoice_payment_details.sales_order_invoice_payment_details_id)
                LEFT JOIN sales_order_payment ON (sales_order_invoice_payment_details.sales_order_payment_id = sales_order_payment.id)
            WHERE sales_order.status_id != 3
                GROUP BY
                    customer.id,
                    customer.address, 
                    customer.contact_person,
                    customer.name,
                    customer.no_telp,
                    customer.daerah,
                    master_list.name,
                    sales_order.id,
                    sales_order.date,
                    sales_order.delivery_date,
                    sales_order_invoice.id,
                    sales_order_invoice.date,
                    sales_order_invoice.due_date,
                    sales_order_invoice.date_completed,
                    sales_order_invoice.total,
                    sales_order.status_id,
                    approval_status.name,
                    payment_method.name, 
                    sales_order_invoice.invoice_number,
                    customer.reference_code
        )
        SELECT
            customer_id,
            customer_address, 
            customer_contact_person,
            customer_name,
            customer_no_telp,
            customer_daerah,
            master_list_name,
            sales_order_id,
            sales_order_date,
            sales_order_delivery_date,
            sales_order_invoice_id,
            sales_order_invoice_date,
            sales_order_invoice_due_date,
            sales_order_invoice_date_completed,
            sales_order_invoice_total,
            sales_order_status_id,
            approval_status_name,
            payment_method_name, 
            sales_order_invoice_invoice_number,
            customer_reference_code,
            total_payment_amount
        FROM RankedSalesOrders
        WHERE rn = 1
        ORDER BY
            master_list_name, 
            customer_reference_code,
            customer_address,
            customer_name,
            sales_order_id
            """
        # Execute a query
        cur.execute(customerQuery)
        # Retrieve query results
        records = cur.fetchall()

        sales_order_ids = [record[CUSTOMERQUERY_SALES_ORDER_ID] for record in records]

        item_query = """
            SELECT 
                sales_order_detail.sales_order_id, 
                sales_order_detail.id, 
                item.name, 
                sales_order_detail.quantity, 
                sales_order_detail.price
            FROM sales_order_detail
            JOIN inventory ON sales_order_detail.inventory_id = inventory.id
            JOIN item ON inventory.item_id = item.id
            WHERE sales_order_detail.sales_order_id = ANY(%s)
            ORDER BY sales_order_detail.sales_order_id
            """

        # Execute the query and fetch the results
        cur.execute(item_query, (sales_order_ids,))
        itemRecords = cur.fetchall()

        debtQuery = """
            SELECT 
                customer.id,
                sales_order.id,
                sales_order.date,
                sales_order.delivery_date,
                sales_order_invoice.id,
                sales_order_invoice.date,
                sales_order_invoice.due_date,
                sales_order_invoice.total,
                payment_method.name,
                sales_order_invoice.invoice_number,
                COALESCE(SUM(sales_order_payment.amount), 0) AS total_payment_amount,
                (sales_order_invoice.total - COALESCE(SUM(sales_order_payment.amount), 0)) AS debt
            FROM
                sales_order
                JOIN customer ON (sales_order.customer_id = customer.id)
                JOIN sales_order_invoice ON (sales_order_invoice.sales_order_id = sales_order.id)
                LEFT JOIN payment_method ON (sales_order_invoice.payment_method_id = payment_method.id)
                LEFT JOIN sales_order_invoice_payment_details ON (sales_order_invoice.id = sales_order_invoice_payment_details.sales_order_invoice_payment_details_id)
                LEFT JOIN sales_order_payment ON (sales_order_invoice_payment_details.sales_order_payment_id = sales_order_payment.id)
            WHERE
                delivery_date IS NOT NULL
                AND date_completed IS NULL
            GROUP BY
                customer.id,
                sales_order.id,
                sales_order_invoice.id,
                payment_method.name,
                sales_order_invoice.invoice_number
            ORDER BY
                customer.name,
                sales_order.id;
            """
        cur.execute(debtQuery)
        debtRecords = cur.fetchall()

        paymentDetailsQuery = """
            SELECT 
                customer.id,
                sales_order.id,
                sales_order_invoice.id,
                sales_order_invoice.date,
                sales_order_payment.date,
                sales_order_payment.amount,
                sales_order_payment.description,
                SUM(sales_order_payment.amount) AS total_payment_amount
            FROM
                (SELECT 
                    customer.id AS customer_id,
                    sales_order.id AS order_id,
                    sales_order_invoice.id AS invoice_id,
                    ROW_NUMBER() OVER (PARTITION BY customer.id ORDER BY sales_order.date DESC) AS row_num
                FROM
                    sales_order
                    JOIN customer ON (sales_order.customer_id = customer.id)
                    JOIN sales_order_invoice ON (sales_order_invoice.sales_order_id = sales_order.id)
                WHERE
                    sales_order.delivery_date IS NOT NULL
                    ) AS subquery    
                JOIN customer ON (subquery.customer_id = customer.id)
                JOIN master_list on (customer.master_list_id = master_list.id)
                JOIN sales_order ON (subquery.order_id = sales_order.id)
                JOIN sales_order_invoice ON (subquery.invoice_id = sales_order_invoice.id)
                LEFT JOIN payment_method ON (sales_order_invoice.payment_method_id = payment_method.id)
                LEFT JOIN sales_order_invoice_payment_details ON (sales_order_invoice.id = sales_order_invoice_payment_details.sales_order_invoice_payment_details_id)
                LEFT JOIN sales_order_payment ON (sales_order_invoice_payment_details.sales_order_payment_id = sales_order_payment.id)
            WHERE
                subquery.row_num <= 1            
                AND amount is not NULL
            GROUP BY
                customer.id,
                sales_order.id,
                sales_order_invoice.id,
                payment_method.name,
                sales_order_payment.date,
                sales_order_payment.amount,
                sales_order_payment.description
            ORDER BY
                customer.id,
                sales_order.id;
        """
        cur.execute(paymentDetailsQuery)
        paymentDetailsRecords = cur.fetchall()

        inactiveCustomerQuery = """
            SELECT 
                customer.id,
                customer.address, 
                customer.contact_person,
                customer.name,
                customer.no_telp,
                customer.daerah,
                master_list.name,
                customer.reference_code
            FROM
                customer 
                LEFT JOIN master_list on (customer.master_list_id = master_list.id)
                LEFT JOIN sales_order on (customer.id = sales_order.customer_id)
            WHERE
                sales_order.id is null
            ORDER BY
                master_list.name, 
                customer.postal_code,
                customer.address,
                customer.name            
            """
        # Execute a query
        cur.execute(inactiveCustomerQuery)
        # Retrieve query results
        inactiveRecords = cur.fetchall()

        customers = []
        for row in records:
            customerCode = str(row[CUSTOMERQUERY_CUSTOMER_ID])
            location = ''
            
            if customerCode in locationDataDict:
                locationData = locationDataDict[customerCode]
                location = locationData[0].get('location')

            customer = customerDto(
            id = row[CUSTOMERQUERY_CUSTOMER_ID],
            address = row[CUSTOMERQUERY_CUSTOMER_ADDRESS],
            contact_person = row[CUSTOMERQUERY_CUSTOMER_CONTACT_PERSON],
            name = row[CUSTOMERQUERY_CUSTOMER_NAME],
            no_telp = row[CUSTOMERQUERY_CUSTOMER_NO_TELP],
            daerah = '',
            master_list=row[CUSTOMERQUERY_MASTER_LIST_NAME],          
            last_complete=None,
            last_duedate=None,
            last_invoice=None,
            last_count_days=None,          
            raport='No Raport',
            last_order='',
            raport_string='No Raport',
            last_delivery='',
            last_payment='',
            debt='',
            payment_details='',
            debt_sort_last=None,
            debt_sort_first=None,
            payment_details_sort_first=None,
            payment_details_sort_last=None,
            master_list_order=row[CUSTOMERQUERY_MASTER_LIST_NAME] + '-99999',
            location=location,
            status_id='No Status'
            )

            if row[CUSTOMERQUERY_CUSTOMER_DAERAH] is not None:
                customer.daerah = row[CUSTOMERQUERY_CUSTOMER_DAERAH]
            if row[CUSTOMERQUERY_SALES_ORDER_DELIVERY_DATE] is not None:
                customer.last_delivery = row[CUSTOMERQUERY_SALES_ORDER_DELIVERY_DATE].strftime('%Y-%m-%d')
            if row[CUSTOMERQUERY_SALES_ORDER_INVOICE_DATE] is not None:
                customer.last_invoice = row[CUSTOMERQUERY_SALES_ORDER_INVOICE_DATE].strftime('%Y-%m-%d')
            if row[CUSTOMERQUERY_SALES_ORDER_INVOICE_DUE_DATE] is not None:
                customer.last_duedate = row[CUSTOMERQUERY_SALES_ORDER_INVOICE_DUE_DATE].strftime('%Y-%m-%d')
            if row[CUSTOMERQUERY_SALES_ORDER_INVOICE_DATE_COMPLETED] is not None:
                customer.last_complete = row[CUSTOMERQUERY_SALES_ORDER_INVOICE_DATE_COMPLETED].strftime('%Y-%m-%d')
            if row[CUSTOMERQUERY_PAYMENT_METHOD_NAME] is not None:
                customer.last_payment = row[CUSTOMERQUERY_PAYMENT_METHOD_NAME]
            if row[CUSTOMERQUERY_CUSTOMER_REFERENCE_CODE] is not None:
                customer.master_list_order = row[CUSTOMERQUERY_MASTER_LIST_NAME] + '-' + row[CUSTOMERQUERY_CUSTOMER_REFERENCE_CODE]
            if row[CUSTOMERQUERY_APPROVAL_STATUS_NAME] is not None:
                customer.status_id = row[CUSTOMERQUERY_APPROVAL_STATUS_NAME]

            lastorderRecords = list(filter(lambda record: record[ITEMQUERY_SALES_ORDER_DETAIL_SALES_ORDER_ID] == row[CUSTOMERQUERY_MASTER_LIST_NAME], itemRecords))
            lastorderString = '';
            if len(lastorderRecords) > 0:
                lastorderString = f"{row[CUSTOMERQUERY_SALES_ORDER_INVOICE_INVOICE_NUMBER]}<br><br>"
                total = 0
                for record in lastorderRecords:
                    lastorderString += f"{record[ITEMQUERY_ITEM_NAME]} <br>({'{:,.0f}'.format(record[ITEMQUERY_SALES_ORDER_DETAIL_QUANTITY])} X {'{:,.0f}'.format(record[ITEMQUERY_SALES_ORDER_DETAIL_PRICE])} = {'{:,.0f}'.format(record[ITEMQUERY_SALES_ORDER_DETAIL_PRICE]*record[ITEMQUERY_SALES_ORDER_DETAIL_QUANTITY])})<br><br>"
                    total = total + (record[ITEMQUERY_SALES_ORDER_DETAIL_PRICE]*record[ITEMQUERY_SALES_ORDER_DETAIL_QUANTITY])
                lastorderString += f"TOTAL {'{:,.0f}'.format(total)}<br><br>"
                customer.last_order = mark_safe(lastorderString)


            oldDebt = False
            currentDebtRecords = list(filter(lambda record: record[DEBTQUERY_CUSTOMER_ID] == row[CUSTOMERQUERY_CUSTOMER_ID], debtRecords))
            sortedDebtRecords = sorted(currentDebtRecords, key=lambda record: record[DEBTQUERY_SALES_ORDER_ID])
            debtString = ''
            if len(sortedDebtRecords) > 0:
                oldDebt = True
                total = 0
                for record in sortedDebtRecords:
                    debtString += f"{record[DEBTQUERY_SALES_ORDER_INVOICE_INVOICE_NUMBER]} <br>{record[DEBTQUERY_SALES_ORDER_DELIVERY_DATE].strftime('%Y-%m-%d')} <br> ({'{:,.0f}'.format(record[DEBTQUERY_SALES_ORDER_INVOICE_TOTAL])} - {'{:,.0f}'.format(record[DEBTQUERY_TOTAL_PAYMENT_AMOUNT])} = {'{:,.0f}'.format(record[DEBTQUERY_DEBT])})<br><br>"
                    total = total + (record[DEBTQUERY_DEBT])
                debtString += f"TOTAL {'{:,.0f}'.format(total)}<br><br>"
                customer.debt_sort_first = sortedDebtRecords[0][DEBTQUERY_SALES_ORDER_DELIVERY_DATE]
                customer.debt_sort_last = sortedDebtRecords[len(sortedDebtRecords)-1][DEBTQUERY_SALES_ORDER_DELIVERY_DATE]
                customer.debt = mark_safe(debtString)
            
            currentPaymentDetailsRecords = list(filter(lambda record: record[PAYMENTDETAILSQUERY_CUSTOMER_ID] == row[CUSTOMERQUERY_CUSTOMER_ID], paymentDetailsRecords))
            sortedPaymentDetailsRecords = sorted(currentPaymentDetailsRecords, key=lambda record: record[PAYMENTDETAILSQUERY_SALES_ORDER_ID])
            paymentDetailsString = '';
            if len(sortedPaymentDetailsRecords) > 0:
                total = 0
                for record in sortedPaymentDetailsRecords:
                    paymentDetailsString += f"{record[PAYMENTDETAILSQUERY_TOTAL_PAYMENT_AMOUNT]} <br>{record[PAYMENTDETAILSQUERY_SALES_ORDER_PAYMENT_DATE].strftime('%Y-%m-%d')} <br> ({'{:,.0f}'.format(record[PAYMENTDETAILSQUERY_SALES_ORDER_PAYMENT_AMOUNT])})<br><br>"
                    total = total + (record[7])
                paymentDetailsString += f"TOTAL {'{:,.0f}'.format(total)}<br><br>"
                customer.payment_details_sort_first = sortedPaymentDetailsRecords[0][PAYMENTDETAILSQUERY_SALES_ORDER_INVOICE_DATE]
                customer.payment_details_sort_last = sortedPaymentDetailsRecords[len(sortedPaymentDetailsRecords)-1][PAYMENTDETAILSQUERY_SALES_ORDER_INVOICE_DATE]
                customer.payment_details = mark_safe(paymentDetailsString)

            dateComplete = None
            dateNunggak = None
            diffInDays = 0

            if customer.last_complete is not None:
                if (oldDebt):
                    customer.raport = "Menunggak Invoice Lama"
                    customer.raport_string = "Menunggak Invoice Lama"
                else:
                    dateComplete = row[CUSTOMERQUERY_SALES_ORDER_INVOICE_DATE_COMPLETED]
                    dateDelivery = row[CUSTOMERQUERY_SALES_ORDER_DELIVERY_DATE].date()
                    diffDate = dateComplete - dateDelivery
                    diffInDays = diffDate.days
                    if diffInDays > 0 :
                        customer.last_count_days = diffInDays
                        customer.raport = "Terlambat"
                        customer.raport_string = "Terlambat " + str(diffInDays) + " hari"
                    elif diffInDays <= 0 :
                        customer.raport = "Lancar"  
                        customer.raport_string = "Lancar"
            elif customer.last_complete is None and customer.last_duedate is not None:
                if (row[CUSTOMERQUERY_SALES_ORDER_DELIVERY_DATE] is not None):
                    dateDelivery = row[CUSTOMERQUERY_SALES_ORDER_DELIVERY_DATE].date()
                    dateNunggak = (datetime.now().date() - dateDelivery).days
                    if (dateNunggak > 0):
                        customer.last_count_days = dateNunggak
                        customer.raport = "Menunggak"
                        customer.raport_string = "Menunggak " + str(dateNunggak) + " hari"
                    else:
                        customer.last_count_days = dateNunggak
                        customer.raport = "In Order"
                        customer.raport_string = "In Order"
                else:
                    customer.raport = "Waiting"  
                    customer.raport_string = "Waiting"

            customers.append(customer)

        for customer in customers:
            if reportType == 'location':              
                row_num += 1          
                data = [customer.master_list_order, customer.id, customer.name, customer.address, customer.last_invoice, customer.last_duedate, customer.last_delivery, customer.last_complete, customer.last_order, customer.last_payment, customer.payment_details, customer.debt, customer.raport, customer.location ]
                for col_num, cell_value in enumerate(data, 1):
                    col_letter = get_column_letter(col_num)
                    worksheet[f"{col_letter}{row_num}"] = cell_value

        # Create a response with the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="data.xlsx"'

        # Save the workbook to the response
        workbook.save(response)

        return response

    else:
        return render(request, 'cnvadmin/reports.html')

def calc_dist(lat_a, long_a, lat_b, long_b):
    # Radius of the Earth in kilometers
    earth_radius_km = 6371.0

    # Convert latitude and longitude from degrees to radians
    lat_a = radians(lat_a)
    long_a = radians(long_a)
    lat_b = radians(lat_b)
    long_b = radians(long_b)

    # Haversine formula to calculate the distance
    dlon = long_b - long_a
    dlat = lat_b - lat_a

    a = sin(dlat / 2)**2 + cos(lat_a) * cos(lat_b) * sin(dlon / 2)**2
    c = 2 * atan2(sqrt(a), sqrt(1 - a))

    # Calculate the distance in kilometers
    distance_km = earth_radius_km * c

    # Convert the distance to meters
    distance_meters = distance_km * 1000.0

    return distance_meters

def custom_sort_key(x):
    if x.customerCode is None or x.customerCode == '':
        # Handle the case where customerCode is None
        customerCode = 9999999
    else:
        customerCode = x.customerCode
    if x.checkInTime is None or x.checkInTime == '':
        # Handle the case where checkInTime is None
        checkInTime = datetime.combine(x.date.replace(tzinfo=None).date(), time(23, 59, 59))
    else:
        checkInTime = x.checkInTime.replace(tzinfo=None)
    return (customerCode, checkInTime)

def custom_sort_key_user(x):
    if x.user_id is None or x.user_id == '':
        # Handle the case where customerCode is None
        user_id = 'wjb@canvasser.com'
    else:
        user_id = x.user_id
    if x.checkInTime is None or x.checkInTime == '':
        # Handle the case where checkInTime is None
        checkInTime = datetime.combine(x.date.replace(tzinfo=None).date(), time(23, 59, 59))
    else:
        checkInTime = x.checkInTime.replace(tzinfo=None)
    return (user_id, checkInTime)